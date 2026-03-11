"""
文档解析与存储
将excel 文件解析为结构化html文件并存储在本地
"""

import json
import os
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import xlrd

# LLM相关库导入
from langchain_openai import ChatOpenAI
from langchain_core.prompts import PromptTemplate
from langchain_core.messages import HumanMessage, SystemMessage


HTML_SCHEMA = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>[文档标题]</title>
    <!-- 可选：基础 viewport meta 标签 -->
    <!-- <meta name="viewport" content="width=device-width, initial-scale=1.0"> -->
</head>
<body>
    <!-- 主容器 -->
    <main class="document">

        <!-- 文档头部信息 -->
        <header class="document__header">
            <h1 class="document__title">[主标题]</h1>
            <p class="document__info-item">[信息项 1: 例如 被审计单位：ABC公司]</p>
            <p class="document__info-item">[信息项 2: 例如 会计期间：2023-01-01至2023-12-31]</p>
            <!-- ... 更多信息项 ... -->
        </header>

        <!-- 工作表 1 (例如 目录 Index) -->
        <section class="worksheet worksheet_type_index">
            <h2 class="worksheet__title">[工作表标题 1: 例如 目录 (索引号: 1370)]</h2>
            <!-- 工作表内的签名/日期信息 -->
            <p class="paragraph paragraph_type_signature">[签名信息: 例如 被审计单位：... 截止日期：... 编制人：张三 2024-1-1  复核人：李四 2024-1-2]</p>

            <!-- 数据表格 -->
            <table class="data-table data-table_type_index">
                <thead class="data-table__header">
                    <tr class="data-table__row">
                        <th class="data-table__header-cell">[表头 1]</th>
                        <th class="data-table__header-cell">[表头 2]</th>
                        <!-- ... 更多表头 ... -->
                    </tr>
                </thead>
                <tbody class="data-table__body">
                    <tr class="data-table__row">
                        <td class="data-table__cell data-table__cell_type_text">[数据 1,1]</td>
                        <td class="data-table__cell data-table__cell_type_text">[数据 1,2]</td>
                        <!-- ... 更多数据 ... -->
                    </tr>
                    <!-- ... 更多数据行 ... -->
                </tbody>
                <!-- 如果有表尾汇总 -->
                <!--
                <tfoot class="data-table__footer">
                    <tr class="data-table__row">
                        <td class="data-table__cell">[表尾数据]</td>
                        ...
                    </tr>
                </tfoot>
                -->
            </table>
        </section>

        <!-- 工作表 2 (例如 明细表 Details) -->
        <section class="worksheet worksheet_type_details">
            <h2 class="worksheet__title">[工作表标题 2: 例如 无形资产明细表 (索引号: 1370-3)]</h2>
            <p class="paragraph paragraph_type_signature">[签名信息]</p>

            <!-- 数据表格 -->
            <table class="data-table data-table_type_main">
                <thead class="data-table__header">
                    <tr class="data-table__row">
                        <th class="data-table__header-cell">[表头 A]</th>
                        <th class="data-table__header-cell">[表头 B]</th>
                        <!-- ... 更多表头 ... -->
                    </tr>
                </thead>
                <tbody class="data-table__body">
                    <tr class="data-table__row">
                        <td class="data-table__cell data-table__cell_type_text">[明细数据 1,A]</td>
                        <td class="data-table__cell data-table__cell_type_number">[明细数据 1,B]</td>
                        <!-- ... 更多明细数据 ... -->
                    </tr>
                    <tr class="data-table__row">
                        <td class="data-table__cell data-table__cell_type_text">[明细数据 2,A]</td>
                        <td class="data-table__cell data-table__cell_type_number">[明细数据 2,B]</td>
                        <!-- ... 更多明细数据 ... -->
                    </tr>
                    <!-- ... 更多明细数据行 ... -->
                </tbody>
            </table>

            <!-- 审计说明段落 -->
            <p class="paragraph paragraph_type_note">审计说明：</p>
            <p class="paragraph">[具体的审计说明内容...]</p>
        </section>

        <!-- 工作表 3 (例如 审定表 Finalization) -->
        <section class="worksheet worksheet_type_finalization">
            <h2 class="worksheet__title">[工作表标题 3: 例如 无形资产审定表 (索引号: 1370-2)]</h2>
            <p class="paragraph paragraph_type_signature">[签名信息]</p>

            <table class="data-table data-table_type_summary">
                <thead class="data-table__header">
                    <tr class="data-table__row">
                        <th class="data-table__header-cell">[表头 X]</th>
                        <th class="data-table__header-cell">[表头 Y]</th>
                        <!-- ... 更多表头 ... -->
                    </tr>
                </thead>
                <tbody class="data-table__body">
                    <tr class="data-table__row">
                        <td class="data-table__cell data-table__cell_type_text">[审定数据 1,X]</td>
                        <td class="data-table__cell data-table__cell_type_number">[审定数据 1,Y]</td>
                        <!-- ... 更多审定数据 ... -->
                    </tr>
                     <!-- ... 更多审定数据行 ... -->
                </tbody>
            </table>
        </section>

        <!-- 工作表 4 (例如 调整分录 Adjustment Entries) -->
        <section class="worksheet worksheet_type_adjustments">
            <h2 class="worksheet__title">[工作表标题 4: 例如 无形资产科目审计调整分录 (索引号: 1370-1)]</h2>
            <p class="paragraph paragraph_type_signature">[签名信息]</p>

            <table class="data-table data-table_type_entries">
                <thead class="data-table__header">
                    <tr class="data-table__row">
                        <th class="data-table__header-cell">[表头 M]</th>
                        <th class="data-table__header-cell">[表头 N]</th>
                        <!-- ... 更多表头 ... -->
                    </tr>
                </thead>
                <tbody class="data-table__body">
                    <tr class="data-table__row">
                        <td class="data-table__cell data-table__cell_type_text">[分录数据 1,M]</td>
                        <td class="data-table__cell data-table__cell_type_text">[分录数据 1,N]</td>
                         <td class="data-table__cell data-table__cell_type_number">[分录数据 1,O]</td>
                        <!-- ... 更多分录数据 ... -->
                    </tr>
                     <!-- ... 更多分录数据行 ... -->
                </tbody>
            </table>
        </section>

        <!-- 更多工作表 ... -->
        <!--
        <section class="worksheet worksheet_type_[specific_type]">
            <h2 class="worksheet__title">[工作表标题 N]</h2>
            <p class="paragraph paragraph_type_signature">[签名信息]</p>
            <table class="data-table data-table_type_[specific_type]">
                 ...
            </table>
            <p class="paragraph paragraph_type_note">审计说明：</p>
            <p class="paragraph">[说明内容...]</p>
        </section>
        -->

    </main>
</body>
</html>
"""

class DocumentParser:
    """
    Excel文档解析器
    将Excel文件解析为指定的JSON结构，并可选地转换为HTML格式
    """
    
    def __init__(self, llm_api_key: str = None, llm_base_url: str = None, llm_model: str = "qwen-plus"):
        """
        初始化文档解析器
        
        Args:
            llm_api_key (str, optional): LLM API密钥
            llm_base_url (str, optional): LLM API基础URL
            llm_model (str): 使用的LLM模型名称
        """
        self.llm = ChatOpenAI(
            model=llm_model,
            api_key=llm_api_key or os.getenv("LLM_API_KEY"),
            base_url=llm_base_url or os.getenv("LLM_BASE_URL"),
            temperature=0.01
        )
    
    def parse_xlsx_to_json(self, file_path: str) -> str:
        """
        解析xlsx文件的信息以及信息内部组织结构，将文件信息转写为json
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            str: JSON格式的字符串，结构为 {"文件名":"", "文件内容":{ "sheet1名称":{}, "sheet2名称":{}, ... }}
        """
        # 检查文件扩展名以确定文件类型
        _, ext = os.path.splitext(file_path)
        ext = ext.lower()
        
        # 获取文件名
        file_name = os.path.basename(file_path)
        
        # 初始化结果结构
        result = {
            "文件名": file_name,
            "文件内容": {}
        }
        
        if ext == '.xlsx':
            # 处理XLSX文件
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                result["文件内容"][sheet_name] = self._parse_sheet_xlsx(sheet)
            
            # 关闭工作簿
            workbook.close()
        elif ext == '.xls':
            # 处理XLS文件
            workbook = xlrd.open_workbook(file_path)
            
            # 遍历所有工作表
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                result["文件内容"][sheet_name] = self._parse_sheet_xls(sheet)
        
        # 返回JSON字符串
        return json.dumps(result, ensure_ascii=False, indent=2)
    
    def _parse_sheet_xlsx(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        解析XLSX单个工作表
        
        Args:
            sheet (Worksheet): 工作表对象
            
        Returns:
            Dict[str, Any]: 工作表数据的字典表示，键为行号，值为该行数据字典（键为列号）
        """
        sheet_data = {}
        
        # 获取最大行数和列数
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        # 遍历所有行和列
        for row_idx in range(1, max_row + 1):
            row_data = {}
            for col_idx in range(1, max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # 处理单元格值，确保可以序列化为JSON
                cell_value = cell.value
                if cell_value is not None:
                    # 处理可能无法序列化为JSON的类型
                    if hasattr(cell_value, 'isoformat'):  # 日期时间类型
                        cell_value = cell_value.isoformat()
                else:
                    cell_value = None
                    
                row_data[col_idx] = cell_value
            
            sheet_data[row_idx] = row_data
        
        return sheet_data
    
    def _parse_sheet_xls(self, sheet) -> Dict[str, Any]:
        """
        解析XLS单个工作表
        
        Args:
            sheet: xlrd工作表对象
            
        Returns:
            Dict[str, Any]: 工作表数据的字典表示，键为行号，值为该行数据字典（键为列号）
        """
        sheet_data = {}
        
        # 获取最大行数和列数
        max_row = sheet.nrows
        max_column = sheet.ncols
        
        # 遍历所有行和列
        for row_idx in range(max_row):
            row_data = {}
            for col_idx in range(max_column):
                cell = sheet.cell(row_idx, col_idx)
                # 处理单元格值，确保可以序列化为JSON
                cell_value = cell.value
                if cell_value is not None:
                    # 处理可能无法序列化为JSON的类型
                    if isinstance(cell_value, (int, float)) and cell.ctype == 3:  # 日期时间类型
                        # 尝试转换为标准日期格式
                        try:
                            cell_value = str(xlrd.xldate.xldate_as_datetime(cell_value, 0))
                        except:
                            pass
                    elif hasattr(cell_value, 'isoformat'):
                        cell_value = cell_value.isoformat()
                else:
                    cell_value = None
                    
                # xlrd索引从0开始，为了保持一致性，我们将其转换为从1开始
                row_data[col_idx + 1] = cell_value
            
            # xlrd索引从0开始，为了保持一致性，我们将其转换为从1开始
            sheet_data[row_idx + 1] = row_data
        
        return sheet_data
    
    def convert_sheet_to_html_section(self, sheet_name: str, sheet_data: Dict[str, Any]) -> str:
        """
        使用LLM将单个工作表数据转换为HTML section
        
        Args:
            sheet_name (str): 工作表名称
            sheet_data (Dict[str, Any]): 工作表数据
            
        Returns:
            str: HTML section代码
        """
        # 构建提示词模板
        prompt_template = PromptTemplate.from_template(
            "你是一个专业的前端开发人员，专门负责将Excel数据转换为结构化的HTML代码。请将以下Excel工作表数据转换为符合HTML_SCHEMA规范的<section>代码片段。\n\n"
            "要求：\n"
            "1. 严格按照HTML_SCHEMA中的结构生成代码\n"
            "2. 只返回<section>部分，不要包含其他内容\n"
            "3. 确保生成的HTML语义化且结构良好\n"
            "4. 为表格数据添加合适的CSS类\n"
            "5. 保持代码缩进一致以提高可读性\n"
            "6. 确保中文内容正确显示\n\n"
            "工作表名称: {sheet_name}\n\n"
            "工作表数据: {sheet_data}\n\n"
            "参考HTML_SCHEMA: {html_schema}\n\n"
            "请只返回<section>标签部分，不要包含其他内容:"
        )
        
        # 使用模板生成提示词
        prompt = prompt_template.format(
            sheet_name=sheet_name,
            sheet_data=json.dumps(sheet_data, ensure_ascii=False, indent=2),
            html_schema=HTML_SCHEMA
        )
        
        # 构造消息
        messages = [
            SystemMessage(content="你是一个专业的前端开发人员，专门负责将Excel数据转换为结构化的HTML代码。"),
            HumanMessage(content=prompt)
        ]
        
        # 调用LLM生成HTML section
        response = self.llm.invoke(messages)
        return response.content
    
    def convert_xlsx_to_html(self, file_path: str, output_path: str = None) -> str:
        """
        将Excel文件转换为完整的HTML文件
        
        Args:
            file_path (str): Excel文件路径
            output_path (str, optional): 输出HTML文件路径，默认为同目录下同名html文件
            
        Returns:
            str: 生成的HTML文件路径
        """
        # 解析Excel为JSON
        json_str = self.parse_xlsx_to_json(file_path)
        json_data = json.loads(json_str)
        
        # 获取文件名（不带扩展名）
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # 默认输出路径
        if output_path is None:
            output_path = os.path.join(os.path.dirname(file_path), f"{base_name}.html")
        
        # 创建HTML主体头部，严格遵守HTML_SCHEMA
        html_content = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
</head>
<body>
    <!-- 主容器 -->
    <main class="document">
        <header class="document__header">
            <h1 class="document__title">{title}</h1>
        </header>
""".format(title=base_name)
        
        # 为每个工作表生成HTML section
        for sheet_name, sheet_data in json_data["文件内容"].items():
            print(f"正在处理工作表: {sheet_name}")
            try:
                section_html = self.convert_sheet_to_html_section(sheet_name, sheet_data)
                html_content += f"\n        {section_html}"
            except Exception as e:
                print(f"处理工作表 '{sheet_name}' 时出错: {e}")
                # 创建一个简单的默认section作为替代
                html_content += f"""
        <section class="worksheet worksheet_type_default">
            <h2 class="worksheet__title">{sheet_name}</h2>
            <p class="paragraph paragraph_type_signature">无法生成结构化内容</p>
            <p class="paragraph">工作表数据解析失败: {str(e)}</p>
        </section>"""
        
        # 结束HTML，严格遵守HTML_SCHEMA
        html_content += """
    </main>
</body>
</html>"""
        
        # 写入文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return output_path
    
if __name__ == '__main__':
    # 测试代码
    excel_file = input('请输入Excel文件路径: ')
    document_parser = DocumentParser(llm_api_key="sk-a36573a3d7c34db492b5cd68d54b5fdd",llm_base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",llm_model="qwen3-coder-flash")
    print(document_parser.convert_xlsx_to_html(excel_file))